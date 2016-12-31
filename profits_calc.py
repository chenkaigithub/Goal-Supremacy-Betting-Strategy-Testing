from openpyxl import Workbook, load_workbook

# ADD YOUR OWN FILE NAMES on the line below
wb = load_workbook("all-euro-data-%d-%d.xlsx" % (2005, 2016), data_only=True)			# data_only=True avoids reading the formulas from inside the cells

ws = wb.active

fair_odds = { -12 : 3.60,			# Fair odds calculated on aprox. 72.000 matches (2005-2016)
-11 : 3.41,
-10 : 3.24,
-9 : 3.08,
-8 : 2.94,
-7 : 2.81,
-6 : 2.69,
-5 : 2.59,
-4 : 2.49,
-3 : 2.39,
-2 : 2.31,
-1 : 2.23,
0 : 2.15,
1 : 2.08,
2 : 2.02,
3 : 1.96,
4 : 1.90,
5 : 1.84,
6 : 1.79,
7 : 1.74,
8 : 1.70,
9 : 1.65,
10 : 1.61,
11 : 1.57,
12 : 1.53 }

last_row = ws.max_row

for n in [12, 6, 4, 2]:						# Starts from Goal Supremacy of -12 to 12 ... down to -2 to 2.

	profits, bets, wins = 0, 0, 0

	for row in range(2, last_row+1):							# Starts iterating through the rows

		goal_sup = int(ws.cell(row = row, column = 14).value)
		
		if -n <= goal_sup <= n:									# Checks games within the desired goal supremacy scores

			win = str(ws.cell(row = row, column = 15).value)
			
			try:
				odds = float(ws.cell(row = row, column = 7).value)		# Skips games that have no Home odds
			except TypeError:
				odds = 0
		
			if odds > fair_odds[goal_sup]:								# If Home odds are higher than the fair odds, bet on game and check outcome

				if win == 'H':

					profits += odds - 1

					wins += 1

				else:

					profits -= 1

				bets += 1

	print("Ratings from %d to %d" % (-n, n))
	print(10*"-")
	print("No of Bets: ", bets)
	print("No of Wins: ", wins)
	print("Profit points: ", profits)
	print("YIELD: ", profits/bets)
	print()





