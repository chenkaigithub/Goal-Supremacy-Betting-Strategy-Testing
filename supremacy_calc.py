from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


'''
Grabs the prepared xlsx files, calculates Goal Supremacy for each match 
and merges it all to a new - massive - xlsx file
'''

# ------------------------------------------ FUNCTIONS --------------------------------------------


# Returns a list with the team's latest 6 game results at the time. Input is team name and the row which it should look backwards of.
def look_back(team, row):

	for rev_row in range(row-1, 1,-1):						# Look at previous rows to find team name

			for rev_col in range(3,5):

				if team == ws.cell(row = rev_row, column = rev_col).value:		# If team is found

					if rev_col == 3:											# Check if it was Home or Away

						played = "H"
					else:

						played = "A"

					team_history = eval(ws.cell(row = rev_row, column = rev_col + 7).value)															# Read team_history from the team's latest game
					
					team_history = [(played, int(ws.cell(row = rev_row, column = 5).value), int(ws.cell(row = rev_row, column = 6).value))] + team_history		# Append latest game to the front of the list e.g. (H, 2, 1)
					
					if len(team_history) > 6:																										# List deletes the oldest game whenever it exceeds a history of 6 games.

						del team_history[-1]

					return team_history

	return []																								# If no match is found (for the first games), return []




# ------------------------------------------ START --------------------------------------------

wb_final = Workbook()

ws_final = wb_final.active						# ws_final is the merge of ALL the sheets

row_final = 2


for z in range(2005, 2016):

	wb = load_workbook("all-euro-data-%d-%d.xlsx" % (z, z+1))

	for ws in wb:

		ws.cell(row= 1, column= 14, value= "Goal Supremacy").font = Font(bold=True)		# Bold column title
		ws.cell(row= 1, column= 13, value= "Away GS").font = Font(bold=True)		# Bold column title
		ws.cell(row= 1, column= 12, value= "Home GS").font = Font(bold=True)		# Bold column title
		ws.cell(row= 1, column= 11, value= "Away History").font = Font(bold=True)		# Bold column title
		ws.cell(row= 1, column= 10, value= "Home History").font = Font(bold=True)		# Bold column title

		# --- Start gathering last 6 games' history ---

		for row in range(2,ws.max_row+1):			# ws.max_row+1 Exclude title AND fist match, which will raise error when trying to look backwards of it.
			
			for col in range(3,5):					# colums of Home and Away teams
				
				print(ws, row, col)
					
				team = ws.cell(row = row, column = col).value

				if team == None:
					break
				
				print(team)
				
				team_history = look_back(team, row)

				ws.cell(row= row, column= col + 7 , value= str(team_history))					# Append latest game to list e.g. (H, 2, 1)


		# --- Start calculating Goal Supremacy points ---

				if len(team_history) == 6:

					points = 0

					for game in team_history:

						if game[0] == "H":

							points += game[1] - game[2]

						elif game[0] == "A":

							points += game[2] - game[1]

					print(points)		
					
					ws.cell(row= row, column= col + 9 , value= points)
			

			try:

				goal_supremacy = int(ws.cell(row= row, column= 12).value) - int(ws.cell(row= row, column= 13).value)
			
				ws.cell(row = row, column = 14, value = goal_supremacy)

			except TypeError:
				pass

			
			if ws.cell(row= row, column= 14).value != None:

				for col in range(1,15):

					ws_final.cell(row = row_final, column = col, value = ws.cell(row= row, column= col).value)

				row_final += 1



for col in range(1,15):																			# Add titles in first row

	ws_final.cell(row= 1, column= col, value = ws.cell(row= 1, column= col).value)

wb_final.save("all-euro-data-2005-2016.xlsx")											# Path and name of newly created file

